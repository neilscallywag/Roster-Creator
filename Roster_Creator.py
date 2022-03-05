from openpyxl import Workbook
import datetime
from datetime import date
from openpyxl.styles import DEFAULT_FONT
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, Color,PatternFill, Border, Side, Alignment, Protection, colors
import inspect

wb =  Workbook()
std=wb['Sheet']
wb.remove(std)

# Variables
DEFAULT_FONT.name = 'Times New Roman'
DEFAULT_FONT.size = 9
starting_month = 3
months = 5
Date_Cell = 'F1'
Days_Cell = 'F2'
ft = Font( name="Times New Roman", size=9, bold=True)
Fill = PatternFill(start_color='000000',end_color='000000',fill_type='solid')
border_style = Side(border_style="thin", color="000000")
border = Border(top=None, left=border_style, right=border_style, bottom=border_style)
name_array = ['TTSH','TMR','aaa']
for month in range(starting_month,starting_month+months+1):
    wb.create_sheet(datetime.date(1900, month, 1).strftime('%b'))
    

for sheet in wb.worksheets:
    
    
    sheet[Date_Cell] = "Date"
    sheet[Days_Cell] = "Day"
    sheet.formula_attributes['G1'] = {'ca':'1'}
    num_days = monthrange(date.today().year, datetime.datetime.strptime(sheet.title, "%b").month)[1]
    sheet['G1'] = '''=DAY(_xlfn.SEQUENCE(1,{},"1-{}-{}",1))'''.format(num_days,sheet.title,date.today().year)
    sheet['G2'] = '''=TEXT(WEEKDAY(_xlfn.SEQUENCE(1,{},"1-{}-{}",1),1),"ddd")'''.format(num_days,sheet.title,date.today().year)
    ft = Font( name="Times New Roman", size=9, bold=True)
    Fill = PatternFill(start_color='000000',
                   end_color='000000',
                   fill_type='solid')
    border_style = Side(border_style="thin", color="000000")

    border = Border(top=None, left=border_style, right=border_style, bottom=border_style)
   
  
    # numdays + 6 because in docs its +1  for some reason

    for row in sheet.iter_rows(min_row=1,max_row=2,min_col=6,max_col=num_days+7):
        for cell in row:
            cell.font = ft
            cell.border = border
            #print(dir(cell))
            cell.alignment = Alignment(horizontal='center')
            sheet.column_dimensions[cell.column_letter].width = 6
            sheet.column_dimensions[cell.column_letter].height = 13.8            
    for row in sheet.iter_rows(min_row=2,min_col=7,max_col=num_days+6):
        for cell in row:
            cell.font = Font(color="FFFFFF",name="Times New Roman", size=9, bold=True)
            cell.fill = Fill
    #create colums for every name in the name_array in column C
    for col in sheet.iter_cols( min_col =3, max_col = 3, max_row=len(name_array)+4):
        #print(col)
        for name in range(len(name_array)):
            for cell in col:
                r =  name+3
                print(cell.column_letter+str(r))
                sheet[cell.column_letter+str(r)]= name_array[name]


    
            
    
    


wb.save('generated.xlsx')

    
